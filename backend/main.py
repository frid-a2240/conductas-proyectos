from fastapi import FastAPI, HTTPException, APIRouter, Depends, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, FileResponse, RedirectResponse
from pydantic import BaseModel
from typing import Optional
from pathlib import Path
from datetime import datetime
from io import BytesIO

from psycopg2.extras import execute_values
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_cursor
from security import (
    hash_password, verify_password, validate_password_strength,
    create_access_token, set_session_cookie, clear_session_cookie,
    require_user, require_admin,
)

app = FastAPI(title="API Conductas-proyectos ISP")

# ─── CORS — orígenes explícitos para que las cookies funcionen en dev ───
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5173",
    ],
    allow_credentials=True,   # ← necesario para cookies httpOnly
    allow_methods=["*"],
    allow_headers=["*"],
)

# Router con prefijo /conductas-proyectos/api — todas las rutas viven aquí
api = APIRouter(prefix="/conductas-proyectos/api")


# ═══════════════════════ Modelos ═══════════════════════
class PreguntaIn(BaseModel):
    texto: str
    valor_id: int

class PreguntaUpdate(BaseModel):
    texto: Optional[str] = None
    valor_id: Optional[int] = None
    activa: Optional[bool] = None

class UsuarioIn(BaseModel):
    nombre: str

class RespuestaIn(BaseModel):
    usuario_id: int
    pregunta_id: int
    respuesta: Optional[int]   # 1-5 (Likert) o None para borrar

class RespuestaBatchItem(BaseModel):
    usuario_id: int
    pregunta_id: int
    respuesta: int   # 0-5 (no acepta None; batch es para envío completo)

class RespuestasBatchIn(BaseModel):
    respuestas: list[RespuestaBatchItem]

class LoginRequest(BaseModel):
    numero_empleado: str
    password: str

class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str


# ═══════════════════════ Guard exclusivo para Cristina (202001) ═══════════════════════
CRISTINA_EMPLEADO = "202001"

def require_cristina(user: dict = Depends(require_user)):
    if user["numero_empleado"] != CRISTINA_EMPLEADO:
        raise HTTPException(status_code=403, detail="No autorizado")
    return user


# ═══════════════════════ AUTH ═══════════════════════
@api.post("/auth/login")
def login(data: LoginRequest, response: Response):
    with get_cursor() as cur:
        cur.execute("""
            SELECT id, numero_empleado, nombre, password_hash, rol,
                   must_change_password, is_active
            FROM auth_users
            WHERE numero_empleado = %s
        """, (data.numero_empleado,))
        row = cur.fetchone()

    if not row or not row["is_active"]:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    if not verify_password(data.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")

    with get_cursor() as cur:
        cur.execute(
            "UPDATE auth_users SET last_login_at = %s WHERE id = %s",
            (datetime.utcnow(), row["id"]),
        )

    token = create_access_token(row["numero_empleado"], row["rol"])
    set_session_cookie(response, token)

    return {
        "user": {
            "numero_empleado": row["numero_empleado"],
            "nombre": row["nombre"],
            "rol": row["rol"],
            "must_change_password": row["must_change_password"],
        },
    }


@api.post("/auth/logout")
def logout(response: Response):
    clear_session_cookie(response)
    return {"ok": True}


@api.get("/auth/me")
def me(user: dict = Depends(require_user)):
    with get_cursor() as cur:
        cur.execute("""
            SELECT numero_empleado, nombre, rol, must_change_password,
                   evaluacion_completada, evaluacion_completada_at
            FROM auth_users WHERE numero_empleado = %s AND is_active = TRUE
        """, (user["numero_empleado"],))
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return row


@api.post("/auth/change-password")
def change_password(data: ChangePasswordRequest, user: dict = Depends(require_user)):
    err = validate_password_strength(data.new_password)
    if err:
        raise HTTPException(status_code=400, detail=err)

    with get_cursor() as cur:
        cur.execute(
            "SELECT password_hash FROM auth_users WHERE numero_empleado = %s",
            (user["numero_empleado"],),
        )
        row = cur.fetchone()
        if not row or not verify_password(data.current_password, row["password_hash"]):
            raise HTTPException(status_code=400, detail="Contraseña actual incorrecta")

        new_hash = hash_password(data.new_password)
        cur.execute("""
            UPDATE auth_users
            SET password_hash = %s,
                must_change_password = FALSE,
                password_changed_at = %s
            WHERE numero_empleado = %s
        """, (new_hash, datetime.utcnow(), user["numero_empleado"]))

    return {"ok": True, "message": "Contraseña actualizada"}


# ═══════════════════════ Valores ═══════════════════════
@api.get("/valores")
def list_valores(user: dict = Depends(require_user)):
    with get_cursor() as cur:
        cur.execute("SELECT * FROM valores ORDER BY orden")
        return cur.fetchall()


# ═══════════════════════ Preguntas ═══════════════════════
@api.get("/preguntas")
def list_preguntas(
    activa: Optional[bool] = None,
    valor_id: Optional[int] = None,
    user: dict = Depends(require_user),
):
    sql = """
        SELECT p.id, p.texto, p.valor_id, p.orden, p.activa,
               v.nombre AS valor_nombre, v.color AS valor_color
        FROM preguntas p
        JOIN valores v ON v.id = p.valor_id
        WHERE 1=1
    """
    params = []
    if activa is not None:
        sql += " AND p.activa = %s"; params.append(activa)
    if valor_id is not None:
        sql += " AND p.valor_id = %s"; params.append(valor_id)
    sql += " ORDER BY p.orden, p.id"
    with get_cursor() as cur:
        cur.execute(sql, params)
        return cur.fetchall()


@api.post("/preguntas")
def create_pregunta(p: PreguntaIn, admin: dict = Depends(require_admin)):
    with get_cursor() as cur:
        cur.execute("SELECT COALESCE(MAX(orden), 0) + 1 AS o FROM preguntas")
        orden = cur.fetchone()["o"]
        cur.execute(
            """INSERT INTO preguntas (texto, valor_id, orden)
               VALUES (%s, %s, %s) RETURNING *""",
            (p.texto, p.valor_id, orden),
        )
        return cur.fetchone()


@api.put("/preguntas/{pid}")
def update_pregunta(pid: int, p: PreguntaUpdate, admin: dict = Depends(require_admin)):
    fields, params = [], []
    for k, v in p.dict(exclude_none=True).items():
        fields.append(f"{k} = %s"); params.append(v)
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    params.append(pid)
    with get_cursor() as cur:
        cur.execute(f"UPDATE preguntas SET {', '.join(fields)} WHERE id = %s RETURNING *", params)
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Pregunta no encontrada")
        return row


@api.delete("/preguntas/{pid}")
def delete_pregunta(pid: int, admin: dict = Depends(require_admin)):
    with get_cursor() as cur:
        cur.execute("DELETE FROM preguntas WHERE id = %s", (pid,))
    return {"ok": True}


# ═══════════════════════ Usuarios (empleados de la matriz) ═══════════════════════
@api.get("/usuarios")
def list_usuarios(user: dict = Depends(require_user)):
    with get_cursor() as cur:
        cur.execute("SELECT * FROM usuarios WHERE activo = TRUE ORDER BY nombre")
        return cur.fetchall()


@api.post("/usuarios")
def create_usuario(u: UsuarioIn, admin: dict = Depends(require_admin)):
    nombre_norm = u.nombre.strip().title()
    with get_cursor() as cur:
        cur.execute(
            "INSERT INTO usuarios (nombre) VALUES (%s) RETURNING *",
            (nombre_norm,),
        )
        return cur.fetchone()


@api.delete("/usuarios/{uid}")
def delete_usuario(uid: int, admin: dict = Depends(require_admin)):
    with get_cursor() as cur:
        cur.execute("UPDATE usuarios SET activo = FALSE WHERE id = %s", (uid,))
    return {"ok": True}


# ═══════════════════════ Respuestas ═══════════════════════
@api.get("/respuestas")
def list_respuestas(
    usuario_id: Optional[int] = None,
    user: dict = Depends(require_user),
):
    """
    Devuelve el PROMEDIO de calificaciones por (usuario_id, pregunta_id),
    considerando a todos los evaluadores que han calificado.
    Formato compatible con el frontend: {usuario_id, pregunta_id, respuesta, n_evaluadores}
    donde `respuesta` es el promedio (float con 2 decimales).
    """
    sql = """
        SELECT usuario_id,
               pregunta_id,
               ROUND(AVG(respuesta)::numeric, 2)::float AS respuesta,
               COUNT(*) AS n_evaluadores
        FROM respuestas
        WHERE respuesta > 0
    """
    params = []
    if usuario_id is not None:
        sql += " AND usuario_id = %s"
        params.append(usuario_id)
    sql += " GROUP BY usuario_id, pregunta_id"
    with get_cursor() as cur:
        cur.execute(sql, params)
        return cur.fetchall()


@api.post("/respuestas")
def upsert_respuesta(r: RespuestaIn, user: dict = Depends(require_user)):
    evaluador = user["numero_empleado"]
    with get_cursor() as cur:
        # Bloqueo si ya envió
        cur.execute(
            "SELECT evaluacion_completada FROM auth_users WHERE numero_empleado = %s",
            (evaluador,),
        )
        row = cur.fetchone()
        if row and row["evaluacion_completada"]:
            raise HTTPException(
                403,
                "Tu evaluación ya fue enviada y no puede modificarse. Contacta a un administrador para reiniciarla."
            )

        # respuesta = None → eliminar el voto (cambio de idea antes de enviar)
        if r.respuesta is None:
            cur.execute(
                """DELETE FROM respuestas
                   WHERE evaluador_numero_empleado = %s
                     AND usuario_id = %s
                     AND pregunta_id = %s""",
                (evaluador, r.usuario_id, r.pregunta_id),
            )
            return {"ok": True, "deleted": True}

        # 0 = sin referencia (se guarda pero no cuenta para promedios)
        if not (0 <= r.respuesta <= 5):
            raise HTTPException(400, "respuesta debe ser un entero entre 0 y 5, o null para borrar")

        cur.execute(
            """INSERT INTO respuestas
                   (evaluador_numero_empleado, usuario_id, pregunta_id, respuesta)
               VALUES (%s, %s, %s, %s)
               ON CONFLICT (evaluador_numero_empleado, usuario_id, pregunta_id)
               DO UPDATE SET respuesta = EXCLUDED.respuesta,
                             fecha_respuesta = CURRENT_TIMESTAMP
               RETURNING *""",
            (evaluador, r.usuario_id, r.pregunta_id, r.respuesta),
        )
        return cur.fetchone()


@api.post("/respuestas/batch")
def upsert_respuestas_batch(data: RespuestasBatchIn, user: dict = Depends(require_user)):
    """Guarda TODAS las respuestas del evaluador en UN solo round-trip.
    Reemplaza N llamadas al POST /respuestas cuando se envía la evaluación."""
    evaluador = user["numero_empleado"]

    with get_cursor() as cur:
        # Bloqueo si ya envió
        cur.execute(
            "SELECT evaluacion_completada FROM auth_users WHERE numero_empleado = %s",
            (evaluador,),
        )
        row = cur.fetchone()
        if row and row["evaluacion_completada"]:
            raise HTTPException(
                403,
                "Tu evaluación ya fue enviada y no puede modificarse. Contacta a un administrador para reiniciarla."
            )

        # Validar rango de cada respuesta
        for r in data.respuestas:
            if not (0 <= r.respuesta <= 5):
                raise HTTPException(
                    400,
                    f"Valor inválido {r.respuesta} para pregunta {r.pregunta_id}: debe estar entre 0 y 5"
                )

        # Upsert en bloque — una sola sentencia SQL
        rows = [
            (evaluador, r.usuario_id, r.pregunta_id, r.respuesta)
            for r in data.respuestas
        ]
        execute_values(
            cur,
            """INSERT INTO respuestas
                   (evaluador_numero_empleado, usuario_id, pregunta_id, respuesta)
               VALUES %s
               ON CONFLICT (evaluador_numero_empleado, usuario_id, pregunta_id)
               DO UPDATE SET respuesta = EXCLUDED.respuesta,
                             fecha_respuesta = CURRENT_TIMESTAMP""",
            rows,
        )

    return {"ok": True, "guardadas": len(data.respuestas)}


# ═══════════════════════ Estado de evaluación ═══════════════════════
@api.get("/evaluacion/estado")
def estado_evaluacion(user: dict = Depends(require_user)):
    """Devuelve el estado del evaluador logueado + progreso."""
    evaluador = user["numero_empleado"]
    with get_cursor() as cur:
        cur.execute("""
            SELECT evaluacion_completada, evaluacion_completada_at
            FROM auth_users WHERE numero_empleado = %s
        """, (evaluador,))
        estado = cur.fetchone() or {"evaluacion_completada": False, "evaluacion_completada_at": None}

        cur.execute("""
            SELECT (SELECT COUNT(*) FROM usuarios  WHERE activo = TRUE) *
                   (SELECT COUNT(*) FROM preguntas WHERE activa = TRUE) AS total
        """)
        total = cur.fetchone()["total"] or 0

        cur.execute("""
            SELECT COUNT(*) AS n FROM respuestas
            WHERE evaluador_numero_empleado = %s
        """, (evaluador,))
        respondidas = cur.fetchone()["n"] or 0

    return {
        **estado,
        "total_celdas": total,
        "celdas_respondidas": respondidas,
    }


@api.post("/evaluacion/completar")
def completar_evaluacion(user: dict = Depends(require_user)):
    """Marca la evaluación del usuario logueado como enviada.
    Valida que TODAS las celdas (usuario activo × pregunta activa) estén llenas."""
    evaluador = user["numero_empleado"]
    with get_cursor() as cur:
        cur.execute(
            "SELECT evaluacion_completada FROM auth_users WHERE numero_empleado = %s",
            (evaluador,),
        )
        row = cur.fetchone()
        if row and row["evaluacion_completada"]:
            raise HTTPException(400, "Tu evaluación ya fue enviada previamente.")

        cur.execute("""
            SELECT (SELECT COUNT(*) FROM usuarios  WHERE activo = TRUE) *
                   (SELECT COUNT(*) FROM preguntas WHERE activa = TRUE) AS total
        """)
        total = cur.fetchone()["total"] or 0

        cur.execute("""
            SELECT COUNT(*) AS n FROM respuestas
            WHERE evaluador_numero_empleado = %s
        """, (evaluador,))
        n = cur.fetchone()["n"] or 0

        if n < total:
            raise HTTPException(
                400,
                f"Debes calificar todas las celdas antes de enviar. Faltan {total - n} de {total}."
            )

        cur.execute("""
            UPDATE auth_users
            SET evaluacion_completada = TRUE,
                evaluacion_completada_at = %s
            WHERE numero_empleado = %s
        """, (datetime.utcnow(), evaluador))

    return {"ok": True, "message": "Evaluación enviada correctamente."}


@api.post("/evaluacion/reiniciar")
def reiniciar_evaluaciones(admin: dict = Depends(require_admin)):
    """(ADMIN) Borra TODAS las respuestas y libera a todos los evaluadores
    para que puedan volver a calificar desde cero."""
    with get_cursor() as cur:
        cur.execute("DELETE FROM respuestas")
        cur.execute("""
            UPDATE auth_users
            SET evaluacion_completada = FALSE,
                evaluacion_completada_at = NULL
        """)
    return {"ok": True, "message": "Evaluaciones reiniciadas. Todos los supervisores pueden evaluar de nuevo."}


# ═══════════════════════ Estatus (solo Cristina — 202001) ═══════════════════════
@api.get("/cristina/estatus")
def cristina_estatus(admin: dict = Depends(require_cristina)):
    """Lista de evaluadores con su estado: Completada o Pendiente.
    Excluye a Sofía (204862) porque es admin de pruebas, no evalúa realmente."""
    with get_cursor() as cur:
        cur.execute("""
            SELECT numero_empleado,
                   nombre,
                   rol,
                   evaluacion_completada,
                   evaluacion_completada_at
            FROM auth_users
            WHERE is_active = TRUE
              AND numero_empleado <> '204862'
            ORDER BY evaluacion_completada DESC, nombre
        """)
        return cur.fetchall()


# ═══════════════════════ Exportar a Excel ═══════════════════════
@api.get("/exportar/excel")
def exportar_excel(admin: dict = Depends(require_admin)):
    with get_cursor() as cur:
        cur.execute("""
            SELECT p.id, p.texto, p.orden, p.valor_id,
                   v.nombre AS valor_nombre, v.color AS valor_color
            FROM preguntas p
            JOIN valores v ON v.id = p.valor_id
            WHERE p.activa = TRUE
            ORDER BY p.orden, p.id
        """)
        preguntas = cur.fetchall()

        cur.execute("SELECT * FROM usuarios WHERE activo = TRUE ORDER BY nombre")
        usuarios = cur.fetchall()

        cur.execute("""
           SELECT usuario_id, pregunta_id,
                  ROUND(AVG(respuesta)::numeric, 2)::float AS respuesta,
                  COUNT(*) AS n_evaluadores
           FROM respuestas
           WHERE respuesta > 0
           GROUP BY usuario_id, pregunta_id
         """)
        respuestas = cur.fetchall()

    mapa = {(r["usuario_id"], r["pregunta_id"]): r["respuesta"] for r in respuestas}

    LIKERT_TXT = {
         1: "Nunca presenta esta conducta",
         2: "Casi nunca presenta esta conducta",
         3: "A veces presenta esta conducta",
         4: "Casi siempre presenta esta conducta",
         5: "Siempre presenta esta conducta",
    }

    LIKERT_FILL = {
        1: "FEE2E2", 2: "FED7AA", 3: "FEF3C7", 4: "D1FAE5", 5: "A7F3D0",
    }
    LIKERT_TEXT_COLOR = {
        1: "991B1B", 2: "9A3412", 3: "92400E", 4: "065F46", 5: "047857",
    }

    wb = Workbook()
    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ━━━ HOJA 1: Matriz ━━━
    ws = wb.active
    ws.title = "Matriz"
    ws.cell(row=1, column=1, value="Empleado").fill = header_fill

    for col, p in enumerate(preguntas, start=2):
        c = ws.cell(row=1, column=col, value=f"P{col-1}")
        c.fill = PatternFill("solid", fgColor=p["valor_color"].lstrip("#"))
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center
        c.border = border
        c2 = ws.cell(row=2, column=col, value=p["texto"])
        c2.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        c2.font = Font(size=9, italic=True, color="6B7280")
        c2.border = border

    c = ws.cell(1, 1)
    c.font = header_font
    c.alignment = center
    c.border = border

    for fila, u in enumerate(usuarios, start=3):
       ws.cell(row=fila, column=1, value=u["nombre"]).border = border
       for col, p in enumerate(preguntas, start=2):
           valor = mapa.get((u["id"], p["id"]))
           c = ws.cell(row=fila, column=col, value=round(valor, 2) if valor else "")
           c.alignment = center
           c.border = border
           if valor and valor > 0:
            nivel = max(1, min(5, round(valor)))  # ← redondea para elegir color
            c.fill = PatternFill("solid", fgColor=LIKERT_FILL[nivel])
            c.font = Font(color=LIKERT_TEXT_COLOR[nivel], bold=True)

    ws.column_dimensions["A"].width = 28
    for i in range(2, 2 + len(preguntas)):
        ws.column_dimensions[get_column_letter(i)].width = 8
    ws.row_dimensions[2].height = 80
    ws.freeze_panes = "B3"

    # ━━━ HOJA 2: Desempeño por Valor ━━━
    ws2 = wb.create_sheet("Desempeño por Valor")
    por_valor = {}
    for p in preguntas:
        v = p["valor_nombre"]
        if v not in por_valor:
            por_valor[v] = {"color": p["valor_color"], "suma": 0, "n": 0, "pend": 0, "total": 0}
        for u in usuarios:
            valor = mapa.get((u["id"], p["id"]))
            por_valor[v]["total"] += 1
            if valor and 1 <= valor <= 5:
                por_valor[v]["suma"] += valor
                por_valor[v]["n"] += 1
            else:
                por_valor[v]["pend"] += 1

    headers = ["Valor", "Respondidas", "Pendientes", "Promedio", "% Cumplimiento"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = center; c.border = border

    for fila, (valor, d) in enumerate(por_valor.items(), start=2):
        ws2.cell(row=fila, column=1, value=valor).border = border
        ws2.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=d["color"].lstrip("#"))
        ws2.cell(row=fila, column=1).font = Font(bold=True, color="FFFFFF")
        ws2.cell(row=fila, column=2, value=d["n"]).border = border
        ws2.cell(row=fila, column=2).alignment = center
        ws2.cell(row=fila, column=3, value=d["pend"]).border = border
        ws2.cell(row=fila, column=3).alignment = center

        promedio = round(d["suma"] / d["n"], 2) if d["n"] else 0
        ws2.cell(row=fila, column=4, value=f"{promedio} / 5.0").border = border
        ws2.cell(row=fila, column=4).alignment = center
        ws2.cell(row=fila, column=4).font = Font(bold=True)

        pct = round((promedio / 5) * 100, 1) if promedio else 0
        c = ws2.cell(row=fila, column=5, value=f"{pct}%")
        c.alignment = center; c.border = border; c.font = Font(bold=True)
        if   pct >= 80: c.fill = PatternFill("solid", fgColor="D1FAE5")
        elif pct >= 50: c.fill = PatternFill("solid", fgColor="FEF3C7")
        else:           c.fill = PatternFill("solid", fgColor="FEE2E2")

    for col, w in enumerate([35, 14, 14, 14, 16], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ━━━ HOJA 3: Detalle por Empleado ━━━
    ws3 = wb.create_sheet("Detalle por Empleado")
    headers = ["Empleado", "Valor", "Pregunta", "Calificación", "Nivel"]
    for col, h in enumerate(headers, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = center; c.border = border

    fila = 2
    for u in usuarios:
        for p in preguntas:
            valor = mapa.get((u["id"], p["id"]))
            ws3.cell(row=fila, column=1, value=u["nombre"]).border = border
            cv = ws3.cell(row=fila, column=2, value=p["valor_nombre"])
            cv.fill = PatternFill("solid", fgColor=p["valor_color"].lstrip("#"))
            cv.font = Font(bold=True, color="FFFFFF", size=10)
            cv.border = border; cv.alignment = left

            ws3.cell(row=fila, column=3, value=p["texto"]).border = border
            ws3.cell(row=fila, column=3).alignment = left

        cal_cell = ws3.cell(row=fila, column=4, value=round(valor, 2) if valor else "—")
        cal_cell.alignment = center; cal_cell.border = border

        nivel_para_texto = max(1, min(5, round(valor))) if valor else None
        nivel_cell = ws3.cell(row=fila, column=5,
                      value=LIKERT_TXT.get(nivel_para_texto, "Sin responder"))
        nivel_cell.alignment = left; nivel_cell.border = border

        if valor and valor > 0:
           nivel = max(1, min(5, round(valor)))
           cal_cell.fill = PatternFill("solid", fgColor=LIKERT_FILL[nivel])
           cal_cell.font = Font(color=LIKERT_TEXT_COLOR[nivel], bold=True)
           nivel_cell.fill = PatternFill("solid", fgColor=LIKERT_FILL[nivel])
           nivel_cell.font = Font(color=LIKERT_TEXT_COLOR[nivel])
        else:
           nivel_cell.font = Font(color="9CA3AF", italic=True)
        fila += 1

    for col, w in enumerate([22, 32, 60, 14, 32], start=1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"conductas_ISP_{fecha}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═══════════════════════ Registrar router ═══════════════════════
# IMPORTANTE: include_router va DESPUÉS de definir todos los endpoints
app.include_router(api)


# ═══════════════════════ Servir el build de React ═══════════════════════
DIST = Path(__file__).parent / "dist"

if DIST.exists():
    app.mount("/conductas-proyectos/assets", StaticFiles(directory=DIST / "assets"), name="assets")

    @app.get("/conductas-proyectos")
    async def redirect_root():
        return RedirectResponse("/conductas-proyectos/")

    @app.get("/conductas-proyectos/{full_path:path}")
    async def serve_spa(full_path: str):
        file_path = DIST / full_path
        if file_path.is_file():
            return FileResponse(file_path)
        return FileResponse(DIST / "index.html")